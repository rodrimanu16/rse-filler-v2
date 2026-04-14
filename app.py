"""
RSE Questionnaire Filler — Saint-Gobain / SGDBF / Point P
FastAPI backend: file upload, question extraction, KA answering, Excel output.
"""

import io
import json
import os
import re
import uuid
import tempfile
import urllib.parse
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.responses import Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

app = FastAPI(title="RSE Questionnaire Filler")

# ─── CONFIG ───────────────────────────────────────────────────────────────────
WORKSPACE_HOST = os.environ.get("DATABRICKS_HOST", "fevm-rodri-tko.cloud.databricks.com")
if not WORKSPACE_HOST.startswith("http"):
    WORKSPACE_HOST = f"https://{WORKSPACE_HOST}"

KA_ENDPOINT = "ka-a2804329-endpoint"
KA_URL = f"{WORKSPACE_HOST}/serving-endpoints/{KA_ENDPOINT}/invocations"

UPLOAD_DIR = Path(tempfile.gettempdir()) / "rse_filler_uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
SAMPLES_DIR = Path(__file__).parent / "samples"

# ─── IN-MEMORY SESSION STORE ──────────────────────────────────────────────────
sessions: dict = {}


# ─── AUTH ─────────────────────────────────────────────────────────────────────
_sdk_client = None


def _get_workspace_client():
    global _sdk_client
    if _sdk_client is None:
        from databricks.sdk import WorkspaceClient
        _sdk_client = WorkspaceClient()
    return _sdk_client


# ─── THEME / FORMAT DETECTION ─────────────────────────────────────────────────
THEME_KEYWORDS: dict[str, list[str]] = {
    "CO2 & Énergie": [
        "co2", "carbone", "émission", "ghg", "scope", "énergie", "renouvelable",
        "climat", "sbt", "mwh", "kwh", "tco2", "gaz à effet", "transition énergétique",
    ],
    "Économie circulaire & Déchets": [
        "déchet", "recyclage", "valorisation", "circulaire", "packaging", "emballage",
        "réemploi", "recup", "biosourcé",
    ],
    "Biodiversité & Environnement": [
        "biodiversité", "eau", "sol", "nature", "natura", "artificialisation",
    ],
    "Santé-Sécurité": [
        "accident", "sécurité", "santé", "tf ", "tg ", "fréquence", "gravité",
        "safefirst", "absentéisme", "rps", "psychosocial", "epi",
    ],
    "Diversité & Inclusion": [
        "femmes", "diversité", "inclusion", "handicap", "égalité", "index", "genre",
        "rqth", "mixité", "rémunération",
    ],
    "Achats responsables": [
        "fournisseur", "achat", "supply", "ecovadis", "chaîne", "approvisionnement",
        "clause", "contrat fournisseur", "esat",
    ],
    "Gouvernance & Éthique": [
        "éthique", "gouvernance", "corruption", "alerte", "conseil", "administrateur",
        "compliance", "conformité", "devoir de vigilance",
    ],
    "Communautés locales": [
        "communauté", "mécénat", "local", "territoire", "emploi", "bâtisseurs",
    ],
}


def detect_theme(text: str) -> str:
    tl = text.lower()
    for theme, kws in THEME_KEYWORDS.items():
        if any(kw in tl for kw in kws):
            return theme
    return "Autre"


def detect_format(text: str) -> str:
    tl = text.lower()
    if any(kw in tl for kw in ["oui / non", "oui/non", "yes/no", "avez-vous", "disposez-vous"]):
        return "yes_no"
    if any(kw in tl for kw in ["%", "pourcentage", "taux", "part de ", "proportion"]):
        return "percent"
    if any(kw in tl for kw in ["combien", "nombre", "tonnes", "mwh", "kwh", "tco2", "effectif", "score /", "€"]):
        return "number"
    return "text"


FORMAT_LABELS = {"yes_no": "Oui / Non", "percent": "Pourcentage", "number": "Numérique", "text": "Texte libre"}


# ─── EXCEL PARSING ────────────────────────────────────────────────────────────
_Q_KW_RE = re.compile(
    r'\b(question|libellé|libelle|intitulé|intitule|indicateur|critère|critere|description)\b',
    re.IGNORECASE,
)
# Strong: "Votre réponse" / "your answer" — wins over "Type de réponse"
_A_STRONG_KW_RE = re.compile(
    r'votre\s+r[ée]ponse|your\s+(answer|response)',
    re.IGNORECASE,
)
# Weak: bare "réponse" / "response" — used only if no strong match found
_A_WEAK_KW_RE = re.compile(
    r'\b(r[ée]ponse|reponse|response|valeur)\b',
    re.IGNORECASE,
)
_T_KW_RE = re.compile(
    r'\b(th[eè]me|cat[eé]gorie|domaine|rubrique|axe)\b',
    re.IGNORECASE,
)
_D_KW_RE = re.compile(
    r'\b(documents?|justificatifs?|sources?|pi[eè]ces?|annexes?|r[eé]f[eé]rences?|commentaires?)\b',
    re.IGNORECASE,
)

# Instruction rows to skip in data extraction
_SKIP_PREFIXES = (
    "envoyé par", "envoye par", "merci de", "note :", "note:", "instructions",
    "veuillez", "prière", "priere", "ce questionnaire", "compléter la colonne",
    "completer la colonne",
)


def _is_header_row(row_vals: list[str]) -> tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
    """Return (q_col, a_col, t_col, d_col) 1-based. Prefer strong answer match over weak."""
    q_col = t_col = d_col = None
    a_col_strong: Optional[int] = None
    a_col_weak:   Optional[int] = None
    for ci, val in enumerate(row_vals):
        v = val.strip()
        if not v:
            continue
        if q_col is None and _Q_KW_RE.search(v):
            q_col = ci + 1
        if t_col is None and _T_KW_RE.search(v):
            t_col = ci + 1
        if d_col is None and _D_KW_RE.search(v):
            d_col = ci + 1
        # Strong always wins; record first weak only if no strong seen yet
        if _A_STRONG_KW_RE.search(v):
            a_col_strong = ci + 1          # keep updating so last strong wins
        elif a_col_weak is None and _A_WEAK_KW_RE.search(v):
            a_col_weak = ci + 1
    a_col = a_col_strong if a_col_strong is not None else a_col_weak
    return q_col, a_col, t_col, d_col


def extract_questions_excel(filepath: Path) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_questions: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row < 2 or max_col < 2:
            continue

        # ── Find header row: scan ALL rows, require BOTH q_col AND a_col ─────
        q_col: Optional[int] = None
        a_col: Optional[int] = None
        t_col: Optional[int] = None
        header_row: Optional[int] = None

        d_col: Optional[int] = None

        for r in range(1, max_row + 1):
            row_vals = [str(ws.cell(row=r, column=c).value or "") for c in range(1, max_col + 1)]
            rq, ra, rt, rd = _is_header_row(row_vals)
            if rq is not None and ra is not None:
                q_col, a_col, t_col, d_col = rq, ra, rt, rd
                header_row = r
                break
            if rq is not None and q_col is None:
                q_col, t_col, d_col = rq, rt, rd
                header_row = r

        if q_col is not None and a_col is None:
            a_col = q_col + 1

        if q_col is None:
            # Fallback: scan every cell for question-like strings
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    val = ws.cell(row=r, column=c).value
                    if not val or not isinstance(val, str):
                        continue
                    text = val.strip()
                    if len(text) < 25:
                        continue
                    tl = text.lower()
                    if any(tl.startswith(p) for p in _SKIP_PREFIXES):
                        continue
                    if not (
                        text.endswith("?")
                        or any(tl.startswith(kw) for kw in
                               ["quel", "comment", "avez", "disposez", "est-ce", "combien",
                                "décrivez", "précisez", "depuis", "votre organisation"])
                    ):
                        continue
                    ans_c = c + 1
                    while ans_c <= max_col and ws.cell(row=r, column=ans_c).value:
                        ans_c += 1
                    all_questions.append({
                        "id": str(uuid.uuid4()),
                        "text": text,
                        "theme": detect_theme(text),
                        "sheet": sheet_name,
                        "row": r,
                        "question_col": get_column_letter(c),
                        "answer_col": get_column_letter(ans_c),
                        "doc_col": None,
                        "answer_format": detect_format(text),
                        "included": True,
                        "answer": None,
                        "confidence": None,
                        "source": None,
                    })
            continue

        # ── Iterate data rows ─────────────────────────────────────────────────
        data_start = (header_row or 0) + 1
        for r in range(data_start, max_row + 1):
            cell_val = ws.cell(row=r, column=q_col).value
            if not cell_val or not isinstance(cell_val, str):
                continue
            text = cell_val.strip()
            if len(text) < 10:
                continue
            # Skip instruction/metadata rows
            if any(text.lower().startswith(p) for p in _SKIP_PREFIXES):
                continue
            # Get theme from dedicated column if present
            if t_col:
                t_val = ws.cell(row=r, column=t_col).value
                theme = str(t_val).strip() if t_val and str(t_val).strip() else detect_theme(text)
            else:
                theme = detect_theme(text)
            ans_val = ws.cell(row=r, column=a_col).value
            all_questions.append({
                "id": str(uuid.uuid4()),
                "text": text,
                "theme": theme,
                "sheet": sheet_name,
                "row": r,
                "question_col": get_column_letter(q_col),
                "answer_col": get_column_letter(a_col),
                "doc_col": get_column_letter(d_col) if d_col else None,
                "answer_format": detect_format(text),
                "included": True,
                "answer": str(ans_val).strip() if ans_val and str(ans_val).strip() else None,
                "confidence": None,
                "source": None,
            })

    return all_questions


# ─── PDF PARSING ──────────────────────────────────────────────────────────────
def extract_questions_pdf(filepath: Path) -> list[dict]:
    try:
        import pdfplumber
    except ImportError:
        raise HTTPException(500, "pdfplumber non installé — PDF non supporté")

    full_text = ""
    try:
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    full_text += t + "\n"
    except Exception as e:
        raise HTTPException(400, f"Erreur lecture PDF: {e}")

    questions: list[dict] = []
    lines = [l.strip() for l in full_text.split("\n") if l.strip()]

    for line in lines:
        is_q = False
        text = line

        # Numbered: "1." "1)" "Q1:"
        m = re.match(r"^[\dA-Z]+[.)]\s+(.{20,})", line)
        if m:
            is_q = True
            text = m.group(1)
        elif len(line) > 25 and line.endswith("?"):
            is_q = True
        elif len(line) > 25 and any(
            line.lower().startswith(kw)
            for kw in ["avez-vous", "disposez-vous", "quel ", "quelle ", "combien",
                       "comment ", "décrivez", "précisez", "depuis quelle"]
        ):
            is_q = True

        if is_q:
            questions.append({
                "id": str(uuid.uuid4()),
                "text": text,
                "theme": detect_theme(text),
                "sheet": "PDF",
                "row": len(questions) + 1,
                "question_col": "A",
                "answer_col": "B",
                "answer_format": detect_format(text),
                "included": True,
                "answer": None,
                "confidence": None,
                "source": None,
            })

    return questions


# ─── KA QUERY ─────────────────────────────────────────────────────────────────
def _clean_display_name(filename: str) -> str:
    """Turn 'Rapport_RSE_SG_2023.pdf' → 'Rapport RSE SG 2023'."""
    name = filename.rsplit(".", 1)[0] if "." in filename else filename
    return name.replace("_", " ").replace("-", " ").strip()


def _parse_citation_url(url: str) -> tuple[str, str]:
    """From a KA citation URL return (volume_path, text_excerpt).

    KA returns URLs like:
      http://localhost:30003/ajax-api/2.0/fs/files/Volumes/cat/sch/vol/file.pdf#:~:text=...
    We extract the /Volumes/... path and decode the text fragment.
    """
    volume_path = ""
    excerpt = ""
    try:
        parsed = urllib.parse.urlparse(url)
        # Extract /Volumes/... from the URL path
        path = parsed.path
        idx = path.find("/Volumes/")
        if idx >= 0:
            volume_path = path[idx:]
        # Decode text fragment from #:~:text=...
        fragment = parsed.fragment  # ":~:text=..."
        if fragment.startswith(":~:text="):
            raw = fragment[len(":~:text="):].split(",")[0]
            text = urllib.parse.unquote_plus(raw).replace("\n", " ").strip()
            excerpt = text[:250]
    except Exception:
        pass
    return volume_path, excerpt


def query_ka(question_text: str, perimetre: str, conversation_id: str) -> dict:
    prompt = (
        f"[Questionnaire RSE — Périmètre : {perimetre}]\n\n"
        f"{question_text}\n\n"
        "Fournis une réponse factuelle et concise, en citant les chiffres clés disponibles "
        "(année 2023 de préférence). Si la donnée n'est pas disponible dans la base, "
        "indique-le clairement plutôt que d'inventer."
    )
    payload = {
        "input": [{"role": "user", "content": prompt}],
        "databricks_options": {"conversation_id": conversation_id},
    }
    try:
        w = _get_workspace_client()
        data = w.api_client.do(
            "POST",
            f"/serving-endpoints/{KA_ENDPOINT}/invocations",
            body=payload,
        )
        output = data.get("output") or []

        references: list[dict] = []
        ref_by_file: dict[str, int] = {}  # canonical key → ref number

        # ── Pass 1: file_search_call items (rich metadata, may not be present) ─
        for item in (output if isinstance(output, list) else []):
            if not isinstance(item, dict) or item.get("type") != "file_search_call":
                continue
            for res in (item.get("output") or item.get("results") or []):
                if not isinstance(res, dict):
                    continue
                fname = res.get("filename") or res.get("title") or ""
                fid   = res.get("file_id") or fname
                key   = fid or fname
                if not key or key in ref_by_file:
                    continue
                num = len(references) + 1
                ref_by_file[key] = num
                if fname and fname != key:
                    ref_by_file[fname] = num

                attrs    = res.get("attributes") or res.get("metadata") or {}
                page_raw = (attrs.get("page") or attrs.get("page_number") or
                            res.get("page") or res.get("page_number"))
                try:   page = int(page_raw) if page_raw is not None else None
                except (ValueError, TypeError): page = None

                score_raw = res.get("score")
                try:   score = round(float(score_raw), 3) if score_raw is not None else None
                except (ValueError, TypeError): score = None

                excerpt = (res.get("text") or res.get("content") or "")[:300].strip()
                references.append({
                    "num": num, "filename": fname,
                    "display_name": _clean_display_name(fname) if fname else "",
                    "page": page, "score": score, "excerpt": excerpt,
                })

        # ── Pass 2: message blocks — concatenate text, handle all annotation types ─
        # KA returns `url_citation` (block-level) or `file_citation` (char-level).
        # For url_citation: annotation covers the whole block → append [N] at block end.
        # For file_citation: use `index` field for precise insertion.
        text_parts: list[str] = []
        char_ann_positions: list[tuple[int, int]] = []  # (char_offset, ref_num)
        running_pos: int = 0

        for item in (output if isinstance(output, list) else []):
            if not isinstance(item, dict) or item.get("type") != "message":
                continue
            for block in (item.get("content") or []):
                if not isinstance(block, dict) or block.get("type") != "output_text":
                    continue
                block_text = block.get("text", "")
                block_ref_nums: list[int] = []

                for ann in (block.get("annotations") or []):
                    if not isinstance(ann, dict):
                        continue
                    atype = ann.get("type")

                    if atype == "url_citation":
                        # Block-level: title holds filename, url holds volume path + excerpt
                        fname    = ann.get("title") or ""
                        ann_url  = ann.get("url") or ""
                        key      = fname
                        if not key:
                            continue
                        if key not in ref_by_file:
                            num = len(references) + 1
                            ref_by_file[key] = num
                            doc_path, excerpt = _parse_citation_url(ann_url)
                            references.append({
                                "num": num, "filename": fname,
                                "display_name": _clean_display_name(fname),
                                "doc_path": doc_path,
                                "page": None, "score": None, "excerpt": excerpt,
                            })
                        rnum = ref_by_file[key]
                        if rnum not in block_ref_nums:
                            block_ref_nums.append(rnum)

                    elif atype in ("file_citation", "file_path"):
                        # Character-level: insert at precise offset
                        fname = ann.get("filename") or ""
                        fid   = ann.get("file_id") or fname
                        idx   = ann.get("index")
                        key   = fid or fname
                        if not key:
                            continue
                        if key not in ref_by_file:
                            num = len(references) + 1
                            ref_by_file[key] = num
                            if fname and fname != key:
                                ref_by_file[fname] = num
                            references.append({
                                "num": num, "filename": fname,
                                "display_name": _clean_display_name(fname),
                                "page": None, "score": None, "excerpt": "",
                            })
                        rnum = ref_by_file.get(key) or ref_by_file.get(fname)
                        if rnum is not None and idx is not None:
                            char_ann_positions.append((running_pos + int(idx), rnum))

                # For url_citation: append markers right after block text
                if block_ref_nums:
                    block_text = block_text + "".join(f"[{n}]" for n in block_ref_nums)

                text_parts.append(block_text)
                running_pos += len(block_text)

        content = "".join(text_parts)

        # Insert file_citation markers (descending to preserve offsets)
        if char_ann_positions:
            for idx, num in sorted(set(char_ann_positions), key=lambda x: x[0], reverse=True):
                idx = min(max(idx, 0), len(content))
                content = content[:idx] + f"[{num}]" + content[idx:]

        # OpenAI-compat fallback
        if not content:
            choices = data.get("choices") or []
            if choices:
                content = choices[0].get("message", {}).get("content", "")

        if not content:
            return {"answer": "", "confidence": 0.3, "source": None,
                    "references": [], "error": f"KA empty response: {str(data)[:200]}"}

        # ── Confidence heuristic ──────────────────────────────────────────────
        source_str = " | ".join(r["filename"] for r in references) if references else "Base de connaissances RSE"
        has_numbers = bool(re.search(r"\d+[\.,]?\d*\s*(%|tCO2|MWh|€|tonnes|/100)", content))
        has_qualifier = any(kw in content.lower() for kw in
                            ["information non disponible", "pas d'information", "n'est pas précisé",
                             "je ne dispose pas", "données non disponibles"])
        if has_qualifier:      confidence = 0.4
        elif has_numbers:      confidence = 0.88
        elif len(content) > 120: confidence = 0.72
        else:                  confidence = 0.55

        return {
            "answer":     content,
            "confidence": confidence,
            "source":     source_str,
            "references": references,
        }
    except Exception as e:
        return {"answer": "", "confidence": 0.0, "error": str(e),
                "source": None, "references": []}


# ─── OUTPUT GENERATION ────────────────────────────────────────────────────────
SG_BLUE  = "00205C"
SG_GREEN = "008751"
WHITE    = "FFFFFF"

def _thin(color="CBD5E1"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def generate_output_excel(session: dict) -> bytes:
    questions   = session["questions"]
    perimetre   = session.get("perimetre", "SGDBF")
    src_name    = session.get("filename", "questionnaire")
    file_type   = session.get("file_type", "pdf")
    orig_path   = session.get("original_path")

    answered = [q for q in questions if q.get("included") and q.get("answer")]

    if file_type == "xlsx" and orig_path and Path(orig_path).exists():
        return _fill_original_excel(orig_path, answered)
    else:
        return _build_qa_excel(answered, perimetre, src_name)


def _fill_original_excel(orig_path: str, answered: list) -> bytes:
    """Write answers back into the original spreadsheet."""
    wb = openpyxl.load_workbook(orig_path)

    ans_fill  = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    ans_font  = Font(color="166534", size=10, name="Calibri")
    ans_align = Alignment(wrap_text=True, vertical="top")

    src_fill  = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    src_font  = Font(color="1D4ED8", size=8, italic=True, name="Calibri")
    src_align = Alignment(wrap_text=True, vertical="top")

    for q in answered:
        sname = q.get("sheet")
        row   = q.get("row")
        acol  = q.get("answer_col")
        dcol  = q.get("doc_col")
        if not (sname and row and acol and sname in wb.sheetnames):
            continue
        try:
            cell = wb[sname][f"{acol}{row}"]
            cell.value     = q["answer"]
            cell.fill      = ans_fill
            cell.font      = ans_font
            cell.alignment = ans_align
        except Exception:
            pass
        # Write structured references to the doc column (page-aware)
        refs = q.get("references") or []
        if dcol and refs:
            parts = []
            for r in refs:
                name = r.get("display_name") or r.get("filename") or ""
                if r.get("page"):
                    name += f" (p. {r['page']})"
                if name:
                    parts.append(f"[{r['num']}] {name}")
            ref_str = "\n".join(parts)
            if ref_str:
                try:
                    sc = wb[sname][f"{dcol}{row}"]
                    sc.value     = ref_str
                    sc.fill      = src_fill
                    sc.font      = src_font
                    sc.alignment = src_align
                except Exception:
                    pass
        elif dcol and q.get("source") and q["source"] != "Base de connaissances RSE":
            try:
                sc = wb[sname][f"{dcol}{row}"]
                sc.value     = q["source"]
                sc.fill      = src_fill
                sc.font      = src_font
                sc.alignment = src_align
            except Exception:
                pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_qa_excel(answered: list, perimetre: str, src_name: str) -> bytes:
    """Build a new Q&A Excel (for PDF inputs or fallback)."""
    wb = openpyxl.Workbook()

    # Group by theme
    by_theme: dict[str, list] = {}
    for q in answered:
        theme = q.get("theme", "Autre")
        by_theme.setdefault(theme, []).append(q)

    first = True
    for theme, qs in by_theme.items():
        ws = wb.active if first else wb.create_sheet()
        ws.title = theme[:31]
        first = False

        _write_qa_sheet(ws, qs, theme, perimetre)

    # Metadata sheet
    meta = wb.create_sheet("À propos")
    _write_meta_sheet(meta, src_name, perimetre, len(answered))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_qa_sheet(ws, qs: list, theme: str, perimetre: str):
    hdrs   = ["#", "Question", "Périmètre", "Format", "Réponse générée par IA", "Documents justificatifs", "Confiance", "Validé ✓"]
    widths = [5, 65, 20, 15, 70, 40, 13, 11]

    hdr_fill  = PatternFill(start_color=SG_BLUE, end_color=SG_BLUE, fill_type="solid")
    hdr_font  = Font(color=WHITE, bold=True, size=10, name="Calibri")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Theme label row
    ws.merge_cells(f"A1:{get_column_letter(len(hdrs))}1")
    tc = ws["A1"]
    tc.value     = f"  {theme}"
    tc.font      = Font(bold=True, size=11, color=WHITE, name="Calibri")
    tc.fill      = PatternFill(start_color=SG_GREEN, end_color=SG_GREEN, fill_type="solid")
    tc.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26

    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = _thin()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    for i, q in enumerate(qs):
        r   = i + 3
        conf = q.get("confidence") or 0.0

        if conf >= 0.80:
            bg, fg, label = "DCFCE7", "166534", f"✓ {int(conf*100)}%"
        elif conf >= 0.55:
            bg, fg, label = "FEF9C3", "854D0E", f"~ {int(conf*100)}%"
        else:
            bg, fg, label = "FEE2E2", "991B1B", f"! {int(conf*100)}%"

        row_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        refs = q.get("references") or []
        if refs:
            ref_str = "\n".join(
                f"[{r['num']}] {r.get('display_name') or r.get('filename','')}"
                + (f" (p. {r['page']})" if r.get("page") else "")
                for r in refs
            )
        elif q.get("source") and q["source"] != "Base de connaissances RSE":
            ref_str = q["source"]
        else:
            ref_str = ""
        vals = [i + 1, q.get("text",""), perimetre,
                FORMAT_LABELS.get(q.get("answer_format","text"), "Texte"),
                q.get("answer",""), ref_str, label, ""]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill      = row_fill
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border    = _thin()
            if ci == 7:
                c.font = Font(color=fg, bold=True, size=10, name="Calibri")
            if ci == 6 and ref_str:
                c.font = Font(color="1D4ED8", size=8, italic=True, name="Calibri")
        ws.row_dimensions[r].height = 65


def _write_meta_sheet(ws, src_name: str, perimetre: str, count: int):
    import datetime
    ws["A1"] = "Questionnaire RSE — Réponses générées par IA"
    ws["A1"].font = Font(bold=True, size=13, color=SG_BLUE, name="Calibri")
    ws.row_dimensions[1].height = 28

    rows = [
        ("Périmètre de réponse", perimetre),
        ("Source questionnaire", src_name),
        ("Questions traitées", count),
        ("Date de génération", datetime.datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Outil", "Assistant RSE — Saint-Gobain / Databricks"),
        ("", ""),
        ("⚠️  Important",
         "Ces réponses ont été générées automatiquement à partir de la base de connaissances RSE. "
         "Elles doivent être relues et validées avant tout envoi à un partenaire."),
    ]
    for r, (k, v) in enumerate(rows, 3):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=10, name="Calibri")
        vc = ws.cell(row=r, column=2, value=v)
        vc.font = Font(size=10, name="Calibri")
        vc.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 18

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 80


# ─── API ROUTES ───────────────────────────────────────────────────────────────

@app.get("/api/debug-ka")
def debug_ka():
    """Diagnostic endpoint — shows env, SDK init, and KA test call."""
    import os
    info: dict = {
        "DATABRICKS_HOST_set":          bool(os.environ.get("DATABRICKS_HOST")),
        "DATABRICKS_TOKEN_set":         bool(os.environ.get("DATABRICKS_TOKEN")),
        "DATABRICKS_CLIENT_ID_set":     bool(os.environ.get("DATABRICKS_CLIENT_ID")),
        "DATABRICKS_CLIENT_SECRET_set": bool(os.environ.get("DATABRICKS_CLIENT_SECRET")),
        "WORKSPACE_HOST":               WORKSPACE_HOST,
        "KA_URL":                       KA_URL,
    }
    # SDK init
    try:
        w = _get_workspace_client()
        info["sdk_host"]  = w.config.host
        info["sdk_init"]  = "ok"
        info["auth_type"] = w.config.auth_type
    except Exception as e:
        info["sdk_init_error"] = str(e)
        return info

    # Current user
    try:
        me = w.current_user.me()
        info["current_user"] = me.user_name
    except Exception as e:
        info["current_user_error"] = str(e)

    # KA call — full raw response so we can inspect the citation structure
    try:
        data = w.api_client.do(
            "POST",
            f"/serving-endpoints/{KA_ENDPOINT}/invocations",
            body={"input": [{"role": "user", "content": "Quel est le score EcoVadis de Saint-Gobain ?"}],
                  "databricks_options": {"conversation_id": "debug-002"}},
        )
        info["ka_status"]        = "ok"
        info["ka_output_types"]  = [item.get("type") for item in (data.get("output") or [])]
        # Summarise annotation types found across all blocks
        ann_types_seen = []
        for item in (data.get("output") or []):
            if item.get("type") == "message":
                for block in (item.get("content") or []):
                    for ann in (block.get("annotations") or []):
                        ann_types_seen.append(ann.get("type"))
        info["annotation_types_seen"] = list(set(ann_types_seen))
        # Test via our parsed query_ka function
        parsed = query_ka("Quel est le score EcoVadis de Saint-Gobain ?", "SGDBF", "debug-003")
        info["parsed_answer_preview"] = parsed.get("answer", "")[:300]
        info["parsed_references_count"] = len(parsed.get("references") or [])
        info["parsed_references"] = parsed.get("references") or []
        info["parsed_source"] = parsed.get("source")
    except Exception as e:
        info["ka_error"] = str(e)

    return info


@app.get("/api/document")
async def get_document(path: str = Query(..., description="UC Volume path: /Volumes/...")):
    """Proxy a Unity Catalog volume file (PDF, etc.) to the browser."""
    if not path.startswith("/Volumes/"):
        raise HTTPException(status_code=400, detail="Path must start with /Volumes/")
    try:
        w = _get_workspace_client()
        download = w.files.download(path)
        filename = path.rsplit("/", 1)[-1]
        mime = "application/pdf" if filename.lower().endswith(".pdf") else "application/octet-stream"
        content = download.contents.read()
        return Response(
            content=content,
            media_type=mime,
            headers={
                "Content-Disposition": f'inline; filename="{filename}"',
                "Cache-Control": "private, max-age=600",
            },
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/samples")
def list_samples():
    if not SAMPLES_DIR.exists():
        return []
    result = []
    for f in sorted(SAMPLES_DIR.glob("*.xlsx")):
        result.append({"name": f.name, "type": "xlsx", "size": f.stat().st_size})
    for f in sorted(SAMPLES_DIR.glob("*.pdf")):
        result.append({"name": f.name, "type": "pdf", "size": f.stat().st_size})
    return result


@app.get("/samples/{filename}")
def get_sample(filename: str):
    path = SAMPLES_DIR / filename
    if not path.exists():
        raise HTTPException(404, "Fichier non trouvé")
    media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return StreamingResponse(
        io.BytesIO(path.read_bytes()), media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    session_id = str(uuid.uuid4())
    session_dir = UPLOAD_DIR / session_id
    session_dir.mkdir(exist_ok=True)

    filename = file.filename or "questionnaire"
    filepath = session_dir / filename
    filepath.write_bytes(await file.read())

    ext = Path(filename).suffix.lower().lstrip(".")
    if ext in ("xlsx", "xls"):
        questions = extract_questions_excel(filepath)
        file_type = "xlsx"
    elif ext == "pdf":
        questions = extract_questions_pdf(filepath)
        file_type = "pdf"
    else:
        raise HTTPException(400, "Format non supporté. Merci d'utiliser Excel (.xlsx) ou PDF.")

    sessions[session_id] = {
        "session_id": session_id,
        "filename": filename,
        "file_type": file_type,
        "original_path": str(filepath),
        "questions": questions,
        "perimetre": "SGDBF",
        "status": "uploaded",
    }

    return {
        "session_id": session_id,
        "filename": filename,
        "file_type": file_type,
        "questions": questions,
        "question_count": len(questions),
    }


class GenerateRequest(BaseModel):
    session_id: str
    perimetre: str
    question_ids: list[str]


@app.post("/api/generate")
def generate_answers(req: GenerateRequest):
    session = sessions.get(req.session_id)
    if not session:
        raise HTTPException(404, "Session non trouvée")

    session["perimetre"] = req.perimetre
    for q in session["questions"]:
        q["included"] = q["id"] in req.question_ids

    to_answer = [q for q in session["questions"] if q["included"]]
    convo_id  = str(uuid.uuid4())

    def event_stream():
        for i, question in enumerate(to_answer):
            result = query_ka(question["text"], req.perimetre, convo_id)
            for q in session["questions"]:
                if q["id"] == question["id"]:
                    q["answer"]     = result.get("answer", "")
                    q["confidence"] = result.get("confidence", 0.5)
                    q["source"]     = result.get("source", "")
                    q["references"] = result.get("references", [])
                    break
            event = {
                "type":        "progress",
                "question_id": question["id"],
                "index":       i + 1,
                "total":       len(to_answer),
                "answer":      result.get("answer", ""),
                "confidence":  result.get("confidence", 0.5),
                "source":      result.get("source", ""),
                "references":  result.get("references", []),
                "error":       result.get("error"),
            }
            yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"

        yield f"data: {json.dumps({'type': 'done', 'session_id': req.session_id})}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream",
                             headers={"X-Accel-Buffering": "no",
                                      "Cache-Control": "no-cache"})


class UpdateAnswerRequest(BaseModel):
    answer: str


class ChatRequest(BaseModel):
    message: str
    conversation_id: str
    perimetre: str = "Groupe Saint-Gobain"


@app.post("/api/chat")
def chat_endpoint(req: ChatRequest):
    result = query_ka(req.message, req.perimetre, req.conversation_id)
    return {
        "answer":     result.get("answer", ""),
        "references": result.get("references", []),
        "confidence": result.get("confidence", 0.5),
        "source":     result.get("source", ""),
        "error":      result.get("error"),
    }


@app.put("/api/answer/{session_id}/{question_id}")
def update_answer(session_id: str, question_id: str, body: UpdateAnswerRequest):
    session = sessions.get(session_id)
    if not session:
        raise HTTPException(404, "Session non trouvée")
    for q in session["questions"]:
        if q["id"] == question_id:
            q["answer"] = body.answer
            return {"ok": True}
    raise HTTPException(404, "Question non trouvée")


@app.get("/api/download/{session_id}")
def download_result(session_id: str):
    session = sessions.get(session_id)
    if not session:
        raise HTTPException(404, "Session non trouvée")
    try:
        data = generate_output_excel(session)
    except Exception as e:
        raise HTTPException(500, f"Erreur génération fichier: {e}")

    stem = Path(session["filename"]).stem
    out_name = f"{stem}_complété_RSE.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{out_name}"'},
    )


# ─── STATIC FRONTEND ─────────────────────────────────────────────────────────
_static = Path(__file__).parent / "static"
if _static.exists():
    app.mount("/", StaticFiles(directory=str(_static), html=True), name="static")

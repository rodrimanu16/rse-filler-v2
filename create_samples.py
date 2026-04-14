"""
Generate sample RSE questionnaire Excel files for demo purposes.
Produces 3 files with different structures to test the parser:
  1. questionnaire_EcoVadis_Carbone_2024.xlsx  — multi-sheet, N°/Question/Format/Périmètre/Réponse
  2. questionnaire_Social_RH_2024.xlsx         — multi-sheet, different column layout
  3. questionnaire_Achats_Responsables.xlsx     — single sheet, Thème column included
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SAMPLES_DIR = Path(__file__).parent / "samples"
SAMPLES_DIR.mkdir(exist_ok=True)

# ─── BRAND COLOURS ────────────────────────────────────────────────────────────
SG_BLUE      = "00205C"
SG_BLUE_MID  = "0047BB"
SG_GREEN     = "008751"
WHITE        = "FFFFFF"
LIGHT_BLUE   = "EEF2FF"
LIGHT_GREEN  = "F0FDF4"
LIGHT_GRAY   = "F8FAFC"
ALT_ROW      = "F1F5F9"
BORDER_COLOR = "CBD5E1"
ANSWER_FILL  = "DCFCE7"
ANSWER_FONT  = "166534"


def thin_border():
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)


def style_header_row(ws, row: int, ncols: int):
    fill = PatternFill(start_color=SG_BLUE, end_color=SG_BLUE, fill_type="solid")
    font = Font(color=WHITE, bold=True, size=10, name="Calibri")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = align
        c.border = thin_border()
    ws.row_dimensions[row].height = 30


def style_data_row(ws, row: int, ncols: int, alt: bool = False):
    bg = ALT_ROW if alt else WHITE
    fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    font = Font(size=10, name="Calibri")
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 55


def style_answer_cell(ws, row: int, col: int):
    """Light green fill for answer cells to make them obvious."""
    c = ws.cell(row=row, column=col)
    c.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")


def add_title_block(ws, title: str, subtitle: str):
    """Merged title row at top."""
    ws.insert_rows(1, 3)
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = title
    t.font = Font(bold=True, size=14, color=SG_BLUE, name="Calibri")
    t.alignment = Alignment(horizontal="left", vertical="center")
    t.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    s = ws["A2"]
    s.value = subtitle
    s.font = Font(size=10, color="64748B", italic=True, name="Calibri")
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8  # spacer


# ══════════════════════════════════════════════════════════════════════════════
# SAMPLE 1 — EcoVadis-style carbon questionnaire (3 sheets)
# Structure: N° | Question | Format attendu | Périmètre | Réponse | Commentaire
# ══════════════════════════════════════════════════════════════════════════════
def make_sample_1():
    wb = openpyxl.Workbook()

    sheets = {
        "Émissions de GES": [
            ("Quelles sont les émissions totales de GES Scope 1 de votre organisation pour l'année 2023 (en tCO2e) ?",
             "Numérique (tCO2e)", "SGDBF"),
            ("Quelles sont les émissions totales de GES Scope 2 (market-based) pour l'année 2023 (en tCO2e) ?",
             "Numérique (tCO2e)", "SGDBF"),
            ("Votre organisation a-t-elle fixé un objectif de réduction des émissions de GES aligné sur la Science Based Targets initiative (SBTi) ?",
             "Oui / Non", "Groupe Saint-Gobain"),
            ("Si oui, quelle est la date cible et le pourcentage de réduction des émissions visé ?",
             "Texte libre (ex : -46 % d'ici 2030)", "Groupe Saint-Gobain"),
            ("Quelle est la part d'énergies renouvelables dans votre mix énergétique global en 2023 (%) ?",
             "Pourcentage (%)", "SGDBF"),
            ("Votre organisation dispose-t-elle d'un système de management de l'énergie certifié ISO 50001 ?",
             "Oui / Non", "SGDBF"),
            ("Quelle est la consommation totale d'énergie de votre organisation en 2023 (en MWh) ?",
             "Numérique (MWh)", "SGDBF"),
            ("Avez-vous mis en place un plan de transition énergétique avec des jalons annuels ?",
             "Oui / Non + description", "SGDBF"),
        ],
        "Déchets & Économie circulaire": [
            ("Quel est le taux global de valorisation des déchets de votre organisation en 2023 (%) ?",
             "Pourcentage (%)", "SGDBF"),
            ("Votre organisation dispose-t-elle d'un programme de collecte et recyclage des déchets de chantier auprès de ses clients ?",
             "Oui / Non", "Point P"),
            ("Combien de tonnes de déchets de chantier ont été collectées via ce programme en 2023 ?",
             "Numérique (tonnes)", "Point P"),
            ("Quelle est la part de matériaux recyclés ou biosourcés dans vos références produits (%) ?",
             "Pourcentage (%)", "SGDBF"),
            ("Avez-vous une politique de réduction des emballages plastiques à usage unique ?",
             "Oui / Non + description", "SGDBF"),
        ],
        "Gouvernance RSE": [
            ("Votre organisation publie-t-elle un rapport RSE ou de développement durable annuel ?",
             "Oui / Non", "Groupe Saint-Gobain"),
            ("Ce rapport est-il vérifié par un organisme tiers indépendant (OTI) ?",
             "Oui / Non + nom de l'OTI", "Groupe Saint-Gobain"),
            ("Quel est le score EcoVadis le plus récent de votre organisation ?",
             "Score /100", "Groupe Saint-Gobain"),
            ("Quelle médaille EcoVadis votre organisation a-t-elle obtenue ?",
             "Bronze / Silver / Gold / Platinum", "Groupe Saint-Gobain"),
            ("Disposez-vous d'un comité ou d'un responsable RSE dédié au sein de la direction ?",
             "Oui / Non", "SGDBF"),
        ],
    }

    col_widths = [6, 72, 24, 28, 36, 30]
    headers = ["N°", "Question", "Format attendu", "Périmètre", "Réponse", "Commentaire"]

    first = True
    for sheet_name, questions in sheets.items():
        ws = wb.active if first else wb.create_sheet()
        ws.title = sheet_name
        first = False

        add_title_block(
            ws,
            f"Questionnaire RSE 2024 — {sheet_name}",
            "Merci de compléter la colonne « Réponse » avec vos données 2023. "
            "Joignez tout document justificatif dans la colonne « Commentaire »."
        )

        # Headers at row 4
        for ci, h in enumerate(headers, 1):
            ws.cell(row=4, column=ci, value=h)
        style_header_row(ws, 4, len(headers))

        for qi, (q_text, fmt, perim) in enumerate(questions):
            r = 5 + qi
            ws.cell(row=r, column=1, value=qi + 1)
            ws.cell(row=r, column=2, value=q_text)
            ws.cell(row=r, column=3, value=fmt)
            ws.cell(row=r, column=4, value=perim)
            ws.cell(row=r, column=5, value="")   # ← answer goes here
            ws.cell(row=r, column=6, value="")
            style_data_row(ws, r, len(headers), alt=(qi % 2 == 1))
            style_answer_cell(ws, r, 5)

        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.freeze_panes = "A5"

    wb.save(SAMPLES_DIR / "questionnaire_EcoVadis_Carbone_2024.xlsx")
    print("✓ questionnaire_EcoVadis_Carbone_2024.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
# SAMPLE 2 — Social / HR questionnaire (2 sheets, different structure)
# Structure: N° | Question | Format attendu | Réponse | Source / Justificatif
# ══════════════════════════════════════════════════════════════════════════════
def make_sample_2():
    wb = openpyxl.Workbook()

    sheets = {
        "Santé & Sécurité au travail": [
            ("Quel est le taux de fréquence des accidents avec arrêt de travail (TF) en 2023 pour votre périmètre France ?",
             "Taux (accidents / million h travaillées)"),
            ("Quel est le taux de gravité des accidents de travail (TG) en 2023 ?",
             "Taux (jours perdus / millier h travaillées)"),
            ("Combien d'accidents mortels ont été enregistrés en 2023 ?",
             "Nombre (0 = objectif)"),
            ("En quoi consiste le programme SafeFirst de Point P et quels en sont les résultats 2023 ?",
             "Description + indicateurs clés"),
            ("Disposez-vous d'un programme de prévention des risques psychosociaux (RPS) formalisé ?",
             "Oui / Non + description sommaire"),
            ("Quel est le taux d'absentéisme global de votre périmètre en 2023 (%) ?",
             "Pourcentage (%)"),
            ("Quelle est la fréquence des exercices de sécurité / formations obligatoires par salarié et par an ?",
             "Nombre d'heures / salarié / an"),
        ],
        "Diversité & Inclusion": [
            ("Quel est l'index de l'égalité professionnelle femmes-hommes (sur 100) de SGDBF pour 2023 ?",
             "Score /100"),
            ("Quelle est la part de femmes dans les effectifs totaux de SGDBF en 2023 (%) ?",
             "Pourcentage (%)"),
            ("Quelle est la part de femmes dans l'encadrement supérieur du Groupe Saint-Gobain (%) ?",
             "Pourcentage (%)"),
            ("Quel est le taux d'emploi de travailleurs reconnus travailleurs handicapés (RQTH) chez Point P (%) ?",
             "Pourcentage (%)"),
            ("Combien de nationalités différentes sont représentées au sein du Groupe Saint-Gobain ?",
             "Nombre"),
            ("Disposez-vous d'une charte de la diversité ou d'un engagement formalisé en matière d'inclusion ?",
             "Oui / Non + lien ou document"),
            ("Quelles actions concrètes avez-vous menées en 2023 pour favoriser l'inclusion des personnes en situation de handicap ?",
             "Description (3 actions max)"),
            ("Quel est l'écart de rémunération entre femmes et hommes pour des postes équivalents (%) ?",
             "Pourcentage (%)"),
        ],
    }

    col_widths = [6, 75, 26, 40, 30]
    headers = ["N°", "Question", "Format attendu", "Réponse", "Source / Justificatif"]

    first = True
    for sheet_name, questions in sheets.items():
        ws = wb.active if first else wb.create_sheet()
        ws.title = sheet_name
        first = False

        add_title_block(
            ws,
            f"Questionnaire RSE 2024 — {sheet_name}",
            "Périmètre de référence : SGDBF / Point P. Données 2023. "
            "Veuillez compléter la colonne « Réponse » et indiquer la source si disponible."
        )

        for ci, h in enumerate(headers, 1):
            ws.cell(row=4, column=ci, value=h)
        style_header_row(ws, 4, len(headers))

        for qi, (q_text, fmt) in enumerate(questions):
            r = 5 + qi
            ws.cell(row=r, column=1, value=qi + 1)
            ws.cell(row=r, column=2, value=q_text)
            ws.cell(row=r, column=3, value=fmt)
            ws.cell(row=r, column=4, value="")   # answer
            ws.cell(row=r, column=5, value="")
            style_data_row(ws, r, len(headers), alt=(qi % 2 == 1))
            style_answer_cell(ws, r, 4)

        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.freeze_panes = "A5"

    wb.save(SAMPLES_DIR / "questionnaire_Social_RH_2024.xlsx")
    print("✓ questionnaire_Social_RH_2024.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
# SAMPLE 3 — Responsible purchasing (single sheet, includes Thème column)
# Structure: N° | Thème | Question | Type de réponse | Votre réponse | Documents
# ══════════════════════════════════════════════════════════════════════════════
def make_sample_3():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Achats Responsables"

    questions = [
        ("Politique achats",
         "Disposez-vous d'une politique d'achats responsables formalisée et approuvée par la direction générale ?",
         "Oui / Non"),
        ("Politique achats",
         "Cette politique couvre-t-elle explicitement les droits humains, les conditions de travail et l'environnement dans la chaîne d'approvisionnement ?",
         "Oui / Non + description"),
        ("Politique achats",
         "Depuis quelle année cette politique est-elle en vigueur ?",
         "Année (ex : 2018)"),
        ("Évaluation fournisseurs",
         "Quel est le score EcoVadis le plus récent de votre organisation (score global sur 100) ?",
         "Score /100"),
        ("Évaluation fournisseurs",
         "Quelle médaille EcoVadis avez-vous obtenue lors de la dernière évaluation ?",
         "Bronze / Silver / Gold / Platinum"),
        ("Évaluation fournisseurs",
         "Quel pourcentage de vos fournisseurs stratégiques ont été évalués sur des critères RSE (EcoVadis ou équivalent) en 2023 ?",
         "Pourcentage (%)"),
        ("Évaluation fournisseurs",
         "Quel score EcoVadis minimum exigez-vous de vos fournisseurs stratégiques ?",
         "Score /100"),
        ("Clauses contractuelles",
         "Intégrez-vous des clauses RSE dans vos contrats fournisseurs (obligations, audits, résiliation) ?",
         "Oui / Non"),
        ("Clauses contractuelles",
         "Quelle procédure appliquez-vous en cas de non-conformité d'un fournisseur aux exigences RSE ?",
         "Description (plan de correction, suspension, résiliation)"),
        ("Économie sociale",
         "Quel est le volume d'achats auprès du secteur adapté (ESAT/EA) en 2023 (€) ?",
         "Montant en euros (€)"),
        ("Économie sociale",
         "Avez-vous un objectif chiffré de développement des achats inclusifs pour 2025 ?",
         "Oui / Non + objectif"),
        ("Reporting",
         "Publiez-vous des données sur la performance RSE de votre chaîne d'approvisionnement dans votre rapport annuel ?",
         "Oui / Non"),
        ("Reporting",
         "Disposez-vous d'un outil de suivi en temps réel de la performance RSE fournisseurs ?",
         "Oui / Non + nom de l'outil"),
    ]

    col_widths = [5, 24, 72, 26, 40, 28]
    headers = ["N°", "Thème", "Question", "Type de réponse", "Votre réponse", "Documents justificatifs"]

    add_title_block(
        ws,
        "Questionnaire Achats Responsables 2024",
        "Envoyé par : [Nom du partenaire]. Merci de compléter la colonne « Votre réponse » avant le [date limite]."
    )

    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    style_header_row(ws, 4, len(headers))

    theme_colors = {
        "Politique achats":     ("EFF6FF", "1D4ED8"),
        "Évaluation fournisseurs": ("F0FDF4", "15803D"),
        "Clauses contractuelles":  ("FFF7ED", "C2410C"),
        "Économie sociale":        ("FAF5FF", "7E22CE"),
        "Reporting":               ("FFF1F2", "BE123C"),
    }

    prev_theme = None
    for qi, (theme, q_text, fmt) in enumerate(questions):
        r = 5 + qi
        ws.cell(row=r, column=1, value=qi + 1)
        # Theme cell — only show label when it changes
        theme_val = theme if theme != prev_theme else ""
        tc = ws.cell(row=r, column=2, value=theme_val)
        if theme != prev_theme:
            bg, fg = theme_colors.get(theme, (WHITE, SG_BLUE))
            tc.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            tc.font = Font(bold=True, size=9, color=fg, name="Calibri")
        prev_theme = theme

        ws.cell(row=r, column=3, value=q_text)
        ws.cell(row=r, column=4, value=fmt)
        ws.cell(row=r, column=5, value="")   # answer
        ws.cell(row=r, column=6, value="")
        style_data_row(ws, r, len(headers), alt=(qi % 2 == 1))
        style_answer_cell(ws, r, 5)

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A5"

    wb.save(SAMPLES_DIR / "questionnaire_Achats_Responsables.xlsx")
    print("✓ questionnaire_Achats_Responsables.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
# SAMPLE 4 — Biodiversité & Eau questionnaire (PDF, plain-text layout)
# Simulates a partner PDF questionnaire to test the PDF parsing flow.
# ══════════════════════════════════════════════════════════════════════════════
def make_sample_pdf():
    import subprocess, tempfile, shutil

    content = """\
Questionnaire RSE 2024 - Biodiversite & Eau
Partenaire : [Nom de votre organisation]
Perimetre : SGDBF / Groupe Saint-Gobain / Point P
Date limite de reponse : [A remplir]

Instructions :
Merci de repondre a chaque question en vous appuyant sur vos donnees 2023.
Joignez tout document justificatif en piece jointe.

----------------------------------------------------------------------
SECTION 1 - BIODIVERSITE
----------------------------------------------------------------------

1. Avez-vous realise un diagnostic de biodiversite sur vos sites principaux en 2023 ?

2. Quelle est la superficie totale des espaces naturels proteges ou geres
   favorablement a la biodiversite sur l'ensemble de vos sites (en m2) ?

3. Avez-vous signe la declaration "Act4nature" ou un engagement equivalent
   en faveur de la biodiversite ?

4. Combien de sites ont mis en place un plan de gestion differenciee
   des espaces verts en 2023 ?

5. Quelles actions concretes avez-vous mises en oeuvre pour reduire
   l'impermeabilisation des sols sur vos sites en 2023 ?

6. Comment votre organisation integre-t-elle les enjeux de biodiversite
   dans sa politique achats (materiaux, bois certifie PEFC/FSC, especes protegees) ?

7. Disposez-vous d'un programme de suivi de la faune et de la flore
   sur vos sites industriels ou logistiques ?

----------------------------------------------------------------------
SECTION 2 - GESTION DE L'EAU
----------------------------------------------------------------------

8. Quel est le volume total d'eau consomme par votre organisation en 2023 (en m3) ?

9. Avez-vous fixe un objectif de reduction de la consommation d'eau
   pour l'horizon 2025 ou 2030 ?

10. Disposez-vous d'un programme de recuperation et de reutilisation
    des eaux de pluie sur vos sites ?

11. Avez-vous identifie vos sites situes en zone de stress hydrique
    selon le WRI Aqueduct ou un referentiel equivalent ?

12. Quel est le taux de recyclage ou de reutilisation des eaux dans
    vos processus industriels en 2023 (%) ?

13. Votre organisation realise-t-elle des rejets dans les milieux
    aquatiques ? Si oui, sont-ils mesures et conformes aux normes en vigueur ?

----------------------------------------------------------------------
Merci de retourner ce questionnaire complete a : rse@votre-partenaire.com
----------------------------------------------------------------------
"""

    # Write plain-text to temp file then convert to PDF via cupsfilter
    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt',
                                    encoding='utf-8', delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    out_path = SAMPLES_DIR / "questionnaire_Biodiversite_2024.pdf"
    result = subprocess.run(
        ["cupsfilter", "-i", "text/plain", "-o", "application/pdf", tmp_path],
        capture_output=True,
    )
    if result.returncode != 0 and not result.stdout:
        raise RuntimeError(f"cupsfilter failed: {result.stderr.decode()[:200]}")

    out_path.write_bytes(result.stdout)
    Path(tmp_path).unlink(missing_ok=True)
    print(f"✓ questionnaire_Biodiversite_2024.pdf ({len(result.stdout):,} bytes)")


if __name__ == "__main__":
    make_sample_1()
    make_sample_2()
    make_sample_3()
    make_sample_pdf()
    print("\nAll sample files created in:", SAMPLES_DIR)
